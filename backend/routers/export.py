"""Export router — generates formatted Excel reports from filtered records."""
from __future__ import annotations

import asyncio
import io
import logging
from datetime import date
from typing import List, Optional

from fastapi import APIRouter
from fastapi.responses import Response
from pydantic import BaseModel

from services.excel_structurer import extract_extra_fields
from store import recommendation_store

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/export", tags=["export"])


class ExportRequest(BaseModel):
    rec_ids: List[str]


# ── openpyxl helpers ──────────────────────────────────────────────────────────

def _make_workbook(rows: list[dict]) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WOM Records"

    # ── Column definitions ──────────────────────────────────────────────────
    columns = [
        # Standard fields
        ("Priority",            18),
        ("Customer",            28),
        ("Sales Order",         16),
        ("Purchase Order",      18),
        ("Job / Project",       20),
        ("Location",            20),
        ("Equipment",           32),
        ("Part Numbers",        36),
        ("Serials",             28),
        ("Certificate Date",    18),
        ("Tested Date",         16),
        ("Recert. Due",         16),
        ("Age (Months)",        14),
        ("Months to Recert",    16),
        ("Status",              18),
        ("Confidence",          14),
        ("Recommendation",      50),
        ("Notes",               30),
        ("Source File",         30),
        # Excel-only extra fields
        ("Document Type",       22),
        ("Issuer",              28),
        ("Address",             32),
        ("Phone",               18),
        ("Fax",                 18),
        ("Serialization",       28),
        ("Applicable Specs",    28),
        ("Authorized Signatory",24),
        ("Signatory Title",     22),
        ("Total Items",         14),
    ]

    # ── Header row style ────────────────────────────────────────────────────
    HEADER_FILL   = PatternFill("solid", fgColor="1B3A6B")
    HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
    EXTRA_FILL    = PatternFill("solid", fgColor="2E6DB4")
    CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    THIN_SIDE     = Side(style="thin", color="DDDDDD")
    THIN_BORDER   = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    # Column header labels
    std_count = 19  # first 19 are standard fields

    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL if col_idx <= std_count else EXTRA_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # ── Priority colour palette ─────────────────────────────────────────────
    PRIORITY_FILLS = {
        "High":          PatternFill("solid", fgColor="FFE0E0"),
        "Medium":        PatternFill("solid", fgColor="FFF3CD"),
        "Low":           PatternFill("solid", fgColor="E8F5E9"),
        "Manual review": PatternFill("solid", fgColor="F3F3F3"),
    }
    STATUS_FILLS = {
        "overdue":      PatternFill("solid", fgColor="FFD6D6"),
        "due-soon":     PatternFill("solid", fgColor="FFF0CC"),
        "active":       PatternFill("solid", fgColor="D6ECD2"),
        "new":          PatternFill("solid", fgColor="D6E8F7"),
    }

    # ── Data rows ───────────────────────────────────────────────────────────
    for r_idx, row in enumerate(rows, start=2):
        priority = row.get("priority") or ""
        status   = (row.get("status") or "").lower()

        row_fill   = PRIORITY_FILLS.get(priority)

        def _fmt_parts(parts) -> str:
            if not parts:
                return ""
            out = []
            for p in parts:
                if isinstance(p, dict):
                    num = p.get("number") or ""
                    desc = p.get("description") or ""
                    qty  = p.get("qty")
                    entry = num
                    if desc:
                        entry += f" – {desc}"
                    if qty:
                        entry += f" (qty {qty})"
                    out.append(entry)
                else:
                    out.append(str(p))
            return "\n".join(out)

        def _fmt_list(items) -> str:
            if not items:
                return ""
            if isinstance(items, list):
                return ", ".join(str(i) for i in items)
            return str(items)

        values = [
            priority,
            row.get("customer") or "",
            row.get("salesOrder") or "",
            row.get("purchaseOrder") or "",
            row.get("jobOrProject") or "",
            row.get("location") or "",
            row.get("equipment") or "",
            _fmt_parts(row.get("partNumbers")),
            _fmt_list(row.get("serials")),
            row.get("certificateDate") or "",
            row.get("testedDate") or "",
            row.get("recertificationDue") or "",
            row.get("ageMonths") or "",
            row.get("monthsToRecert") or "",
            row.get("status") or "",
            row.get("confidence") or "",
            row.get("recommendation") or "",
            row.get("notes") or "",
            row.get("sourceFile") or "",
            # Extra fields
            row.get("_extra", {}).get("documentType") or "",
            row.get("_extra", {}).get("issuer") or "",
            row.get("_extra", {}).get("address") or "",
            row.get("_extra", {}).get("phone") or "",
            row.get("_extra", {}).get("fax") or "",
            row.get("_extra", {}).get("serialization") or "",
            row.get("_extra", {}).get("applicableSpecs") or "",
            row.get("_extra", {}).get("authorizedSignatory") or "",
            row.get("_extra", {}).get("signatoryTitle") or "",
            row.get("_extra", {}).get("totalItems") or "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=value)
            cell.alignment = LEFT
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[r_idx].height = 40 if "\n" in str(values[7]) else 20

    # ── Freeze header + auto-filter ─────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Route ─────────────────────────────────────────────────────────────────────

@router.post("/excel")
async def export_excel(body: ExportRequest):
    """
    Generate a formatted Excel report for the given recommendation IDs.
    Enriches each record with extra CoC fields via Azure OpenAI.
    """
    if not body.rec_ids:
        return Response(content=b"", status_code=400)

    # Fetch records from Firestore (fast, synchronous store)
    recs = []
    for rec_id in body.rec_ids:
        rec = recommendation_store.get(rec_id)
        if rec:
            recs.append(rec)

    if not recs:
        return Response(content=b"", status_code=404)

    # Run OpenAI extra-field extraction in parallel (max 5 concurrent)
    sem = asyncio.Semaphore(5)

    async def _extract(rec):
        async with sem:
            extra = await extract_extra_fields(rec.textPreview)
            d = rec.model_dump()
            d["_extra"] = extra
            return d

    rows = await asyncio.gather(*[_extract(r) for r in recs])

    # Generate Excel bytes
    try:
        excel_bytes = _make_workbook(list(rows))
    except Exception as exc:
        logger.error("Excel generation failed: %s", exc)
        return Response(content=b"", status_code=500)

    today = date.today().isoformat()
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="wom-records-{today}.xlsx"',
            "Content-Length": str(len(excel_bytes)),
        },
    )
